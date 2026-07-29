# -*- coding: utf-8 -*-
"""
报告组装模块
职责：将所有数据组装到 Excel 各 Sheet（概念总排名、详细个股排名、TOP5、板块明细）
扩展接口：继承 ReporterBase 可新增输出格式（如 PDF、HTML）
"""
import os
import json
from abc import ABC, abstractmethod
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .styler import ExcelStyler, V8Style
from .ranker import Top5Ranker, SectorRanker, DetailRanker


def _fmt_stat(val):
    """统计指标显示格式化：None/缺失值显示为 '-'"""
    return "-" if val is None else val


class ReporterBase(ABC):
    """报告生成器基类"""

    @abstractmethod
    def build(self, data: dict, output_path: str):
        """
        生成报告
        :param data: 所有数据字典
        :param output_path: 输出文件路径
        """
        pass


class ExcelReporter(ReporterBase):
    """
    Excel 报告组装器
    将数据写入以下 Sheet：
    1. 1-概念总排名
    2. 2-详细个股排名
    3. 3-TOP5
    4~. 各板块明细（PCB、半导体等）
    """

    # openpyxl Sheet 标题非法字符
    _INVALID_SHEET_CHARS = '[ ] : * ? / \\'

    @staticmethod
    def _sanitize_sheet_title(title: str) -> str:
        """清洗 Sheet 标题，替换非法字符为 -"""
        if not title:
            return "Sheet"
        for ch in ExcelReporter._INVALID_SHEET_CHARS:
            title = title.replace(ch, "-")
        # 限制长度（openpyxl 最大 31 字符）
        return title[:31]

    def __init__(self, config: dict = None, styler: ExcelStyler = None):
        self.config = config or self._load_config()
        self.styler = styler or ExcelStyler(V8Style(
            header_bg=self.config.get("style", {}).get("header_bg", "4472C4"),
            header_font=self.config.get("style", {}).get("header_font", "FFFFFF")
        ))
        self.sheets_cfg = self.config.get("sheets", {})
        self.scoring_cfg = self.config.get("scoring", {})

    @staticmethod
    def _load_config() -> dict:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_path = os.path.join(base_dir, "config.json")
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _write_dataframe(self, ws, df: pd.DataFrame, start_row: int = 1,
                         header_style: bool = True, data_style: bool = True):
        """将 DataFrame 写入工作表，并应用样式"""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row and header_style:
                self.styler.apply_header_style(ws, r_idx, 1, len(row))
            elif data_style:
                self.styler.apply_data_style(ws, r_idx, 1, len(row))
        return start_row + len(df) + 1

    def _build_info_sheet(self, wb, stats: dict = None):
        """
        构建 Sheet 0: 说明页（打分规则 + 执行统计指标）
        放在报告最前面
        """
        ws = wb.create_sheet(title="0-说明", index=0)

        # ========== 第一部分：打分规则说明 ==========
        title_row = 1
        self.styler.merge_and_title(ws, title_row, 1, 4, "=== 打分规则说明 ===")

        rules = [
            ["维度", "名称", "分值", "规则说明"],
            ["D1", "强势形态且新高", "8分", "最高价创近150日新高"],
            ["D2", "强势形态", "6分", "近5日涨幅>20% 且 最高价是近20日最高价"],
            ["D4", "首板资金池", "3分", "当日首板涨停（前日非涨停）"],
            ["D5", "潜在突破10日", "2分", "最高价>近10日最高价"],
            ["D6", "潜在突破5日", "1分", "最高价>近5日最高价 且 当日非涨停（满足D5则不计）"],
            ["D7", "持续性", "5分", "当日二板及以上（连板天数>=2）"],
            ["D8", "情绪分数", "6分", "当日一字板涨停"],
            ["D9", "活跃程度", "3分", "近10日有涨停板"],
            ["+", "大成交额额外加分", "3分", "当日成交额>=50亿"],
            ["", "合计", "37分", "满分 = 8+6+3+2+1+5+6+3+3 = 37分"],
        ]

        light_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for r_idx, row_data in enumerate(rules, 2):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 2:  # 表头行
                    cell.fill = light_green
                    cell.font = self.styler.style.header_font_style
                    cell.alignment = self.styler.style.header_alignment
                else:
                    self.styler.style.apply(ws, r_idx, c_idx, "data")
                cell.border = self.styler.style.thin_border

        # ========== 第二部分：执行统计指标 ==========
        current_row = len(rules) + 3
        self.styler.merge_and_title(ws, current_row, 1, 4, "=== 执行统计指标 ===")
        current_row += 1

        stats_data = [
            ["指标", "结果", "备注", ""],
            ["报告日期", stats.get('report_date', '-'), "命令传入或默认的日期", ""],
            ["实际数据日期", stats.get('actual_data_dates', '-'), "行情数据实际使用的K线日期（如与报告日期不一致请警惕）", ""],
            ["韭研标签标的", f"{stats.get('pool_count', '-')} 只", "带韭研概念标签的打分池数量", ""],
            ["行情数据成功", f"{stats.get('market_success', '-')}/{stats.get('pool_count', '-')} 只",
             (f"{stats.get('market_failed', '-')} 只失败: {', '.join(stats.get('failed_codes', [])[:20])}{'...' if len(stats.get('failed_codes', [])) > 20 else ''}" if stats and stats.get('market_failed') and stats.get('failed_codes') else f"{stats.get('market_failed', '-')} 只失败，已跳过" if stats and stats.get('market_failed') else ""), ""],

            ["概念总排名", f"{stats.get('sector_count', '-')} 个板块", "按韭研分类拆分后的大板块数", ""],
            ["子概念排名", f"{stats.get('detail_count', '-')} 只标的", "按韭研概念细分后的标的人数", ""],
            ["TOP5", f"{stats.get('top5_count', '-')} 只，{stats.get('top5_sectors', '-')} 个板块",
             "多样性通过" if stats and stats.get('top5_diversity_ok') else "", ""],
            ["验证结果", stats.get('validate_result', '-'), "", ""],
            ["评分范围", stats.get('score_range', '-'), f"平均分: {stats.get('score_avg', '-')}", ""],
            ["板块总体平均分", f"加权 {_fmt_stat(stats.get('overall_avg_weighted'))} / 等权 {_fmt_stat(stats.get('overall_avg_simple'))}",
             "各板块平均分按股票数加权；等权为板块简单平均", ""],
            ["板块最高/最低", f"{stats.get('overall_top_sector') or '-'} / {stats.get('overall_bottom_sector') or '-'}",
             "平均分最高/最低的板块（括号内为该板块平均分）", ""],
            ["热度分最高/最低", f"{stats.get('heat_top_sector') or '-'} / {stats.get('heat_bottom_sector') or '-'}",
             "板块热度分（0-100），基于涨停密度+上涨广度+量能放大+前排强度", ""],
        ]

        light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for r_idx, row_data in enumerate(stats_data, current_row):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == current_row:  # 表头行
                    cell.fill = light_blue
                    cell.font = self.styler.style.header_font_style
                    cell.alignment = self.styler.style.header_alignment
                else:
                    self.styler.style.apply(ws, r_idx, c_idx, "data")
                cell.border = self.styler.style.thin_border

        # ========== 第三部分：板块股票分布 ==========
        if stats and stats.get("sector_stocks"):
            current_row = current_row + len(stats_data) + 1
            self.styler.merge_and_title(ws, current_row, 1, 3, "=== 板块股票分布 ===")
            current_row += 1
            sector_rows = [["板块", "股票总数", ""]]
            for name, count in sorted(stats["sector_stocks"].items(), key=lambda x: -x[1]):
                sector_rows.append([name, f"{count} 只", ""])
            # 增加合计行
            sector_rows.append(["合计", f"{sum(stats['sector_stocks'].values())} 只", ""])
            
            light_orange = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
            for r_idx, row_data in enumerate(sector_rows, current_row):
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == current_row or r_idx == current_row + len(sector_rows) - 1:  # 表头或合计行
                        cell.fill = light_orange
                        cell.font = self.styler.style.header_font_style
                        cell.alignment = self.styler.style.header_alignment
                    else:
                        self.styler.style.apply(ws, r_idx, c_idx, "data")
                    cell.border = self.styler.style.thin_border

        self.styler.auto_width(ws)
        return ws

    def _build_summary_sheet(self, wb, df_summary: pd.DataFrame, stats: dict = None):
        """构建 Sheet 1: 概念总排名（P0: 热度可视化增强）"""
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        ws = wb.create_sheet(title=self.sheets_cfg.get("summary", "1-概念总排名"))
        stats = stats or {}
        has_heat = not df_summary.empty and "热度分" in df_summary.columns
        merge_width = max(15, len(df_summary.columns) if not df_summary.empty else 11)

        # ---- Row 1: 标题 ----
        self.styler.merge_and_title(ws, 1, 1, merge_width, "=== 概念板块全局概览 ===")

        # ---- Row 2: 🔥 热度TOP5速览条（P0: 始终显示，修复死代码）----
        if has_heat:
            heat_top5 = df_summary.head(5)
            top5_parts = []
            for _, r in heat_top5.iterrows():
                trend_icon = r.get('趋势', '')
                top5_parts.append(f"{r['板块']}({r['热度分']}{trend_icon})")
            heat_bar = "🔥 热度TOP5: " + " | ".join(top5_parts)
            heat_cell = ws.cell(row=2, column=1, value=heat_bar)
            heat_cell.font = Font(name="微软雅黑", bold=True, color="CC0000", size=11)
            heat_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_width)

        # ---- Row 3: 辅助信息（平均分/热度最高最低）----
        current_info_row = 2
        overview_parts = []
        w_avg = stats.get("overall_avg_weighted")
        s_avg = stats.get("overall_avg_simple")
        if w_avg is not None or s_avg is not None:
            overview_parts.append(f"板块平均分：加权 {_fmt_stat(w_avg)} / 等权 {_fmt_stat(s_avg)}")
        heat_top_s = stats.get("heat_top_sector")
        heat_bottom_s = stats.get("heat_bottom_sector")
        if heat_top_s or heat_bottom_s:
            overview_parts.append(f"热度最高 {heat_top_s or '-'}　最低 {heat_bottom_s or '-'}")
        if overview_parts:
            info_row = 3 if has_heat else 2
            info_txt = "　｜　".join(overview_parts)
            info_cell = ws.cell(row=info_row, column=1, value=info_txt)
            info_cell.font = Font(name="微软雅黑", color="333333", size=9)
            info_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=merge_width)
            current_info_row = info_row

        data_start_row = current_info_row + 1

        # ---- 写入数据表 ----
        next_row = self._write_dataframe(ws, df_summary, start_row=data_start_row)

        # ---- P0: 热度分列色阶（白→黄→红）----
        if has_heat:
            heat_col_idx = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=data_start_row, column=c).value == "热度分":
                    heat_col_idx = c
                    break
            if heat_col_idx:
                heat_col_letter = get_column_letter(heat_col_idx)
                data_end = data_start_row + len(df_summary)
                if data_end > data_start_row:
                    ws.conditional_formatting.add(
                        f"{heat_col_letter}{data_start_row+1}:{heat_col_letter}{data_end}",
                        ColorScaleRule(
                            start_type="min", start_color="FFFFFF",
                            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                            end_type="max", end_color="FF4444"
                        )
                    )

            # ---- P0: 趋势列按标签填底色 ----
            trend_col_idx = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=data_start_row, column=c).value == "趋势":
                    trend_col_idx = c
                    break
            if trend_col_idx:
                trend_fills = {
                    "🔥加速": PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),
                    "📈升温": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
                    "➡️平稳": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
                    "📉退潮": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
                    "🧊冰点": PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid"),
                }
                trend_fonts = {
                    "🔥加速": Font(name="微软雅黑", bold=True, color="FFFFFF", size=9),
                    "🧊冰点": Font(name="微软雅黑", bold=True, color="FFFFFF", size=9),
                }
                for r in range(data_start_row + 1, data_start_row + 1 + len(df_summary)):
                    cell = ws.cell(row=r, column=trend_col_idx)
                    label = str(cell.value or "")
                    fill = trend_fills.get(label)
                    if fill:
                        cell.fill = fill
                    font = trend_fonts.get(label)
                    if font:
                        cell.font = font

            # ---- P0: 热度分 TOP3 行加粗 ----
            for r in range(data_start_row + 1, min(data_start_row + 4, data_start_row + 1 + len(df_summary))):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.font:
                        cell.font = Font(
                            name=cell.font.name or "微软雅黑",
                            bold=True,
                            color=cell.font.color or "000000",
                            size=cell.font.size or 9
                        )

        self.styler.auto_width(ws)
        return ws

    def _build_detail_sheet(self, wb, df_detail: pd.DataFrame):
        """
        构建 Sheet 2: 子概念排名
        按平均分排序，去掉总得分和平均成交额
        """
        ws = wb.create_sheet(title="2-概念排名")

        title_row = 1
        self.styler.merge_and_title(ws, title_row, 1, 10, "=== 子概念排名（按平均分）===")

        # 写入数据
        if not df_detail.empty:
            self._write_dataframe(ws, df_detail, start_row=2)
        else:
            ws.cell(row=2, column=1, value="暂无数据")

        self.styler.auto_width(ws)
        return ws

    def _build_top5_sheet(self, wb, top5_list: List[dict], df_summary: pd.DataFrame = None):
        """构建 Sheet 3: TOP5（P1: 删D3死列，增板块热度分/趋势）"""
        ws = wb.create_sheet(title=self.sheets_cfg.get("top5", "3-TOP5"))

        ncols = 17  # 排名+代码+名称+板块+得分+涨跌幅+成交额+8dim+热度分+趋势
        self.styler.merge_and_title(ws, 1, 1, ncols, "=== 全市场 TOP5 强势标的 ===")

        if not top5_list:
            ws.cell(row=2, column=1, value="暂无数据")
            return ws

        # 构建板块热度查找表
        heat_map = {}
        if df_summary is not None and not df_summary.empty and "板块" in df_summary.columns:
            if "热度分" in df_summary.columns:
                for _, r in df_summary.iterrows():
                    heat_map[str(r["板块"])] = {
                        "热度分": r.get("热度分", 0),
                        "趋势": r.get("趋势", ""),
                    }

        columns = ["排名", "代码", "名称", "所属板块", "总得分",
                   "涨跌幅", "成交额(亿)",
                   "D1强势形态且新高", "D2强势形态", "D4首板资金池",
                   "D5潜在突破10日", "D6潜在突破5日",
                   "D7持续性", "D8情绪分数", "D9活跃程度", "大成交额额外加分",
                   "板块热度分", "板块趋势"]
        rows = []
        for stock in top5_list:
            sector = stock.get("所属板块", stock.get("板块", ""))
            hi = heat_map.get(sector, {})
            rows.append({
                "排名": stock.get("排名", ""),
                "代码": stock.get("代码", ""),
                "名称": stock.get("名称", ""),
                "所属板块": sector,
                "总得分": stock.get("总得分", 0),
                "涨跌幅": stock.get("涨跌幅", 0),
                "成交额(亿)": round(stock.get("成交额", 0) / 100000000, 2) if stock.get("成交额", 0) else 0,
                "D1强势形态且新高": stock.get("D1强势形态且新高", 0),
                "D2强势形态": stock.get("D2强势形态", 0),
                "D4首板资金池": stock.get("D4首板资金池", 0),
                "D5潜在突破10日": stock.get("D5潜在突破10日", 0),
                "D6潜在突破5日": stock.get("D6潜在突破5日", 0),
                "D7持续性": stock.get("D7持续性", 0),
                "D8情绪分数": stock.get("D8情绪分数", 0),
                "D9活跃程度": stock.get("D9活跃程度", 0),
                "大成交额额外加分": stock.get("大成交额额外加分", 0),
                "板块热度分": hi.get("热度分", 0),
                "板块趋势": hi.get("趋势", ""),
            })
        df_top5 = pd.DataFrame(rows, columns=columns)
        self._write_dataframe(ws, df_top5, start_row=2)

        # 板块趋势列着色（复用 summary 的色表）
        trend_fills = {
            "🔥加速": PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid"),
            "📈升温": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
            "➡️平稳": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
            "📉退潮": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
            "🧊冰点": PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid"),
        }
        trend_fonts = {
            "🔥加速": Font(name="微软雅黑", bold=True, color="FFFFFF", size=9),
            "🧊冰点": Font(name="微软雅黑", bold=True, color="FFFFFF", size=9),
        }
        for r in range(3, 3 + len(top5_list)):
            cell = ws.cell(row=r, column=ncols)  # 最后列=趋势
            label = str(cell.value or "")
            fill = trend_fills.get(label)
            if fill:
                cell.fill = fill
            font = trend_fonts.get(label)
            if font:
                cell.font = font

        self.styler.auto_width(ws)
        return ws

    def _build_sector_sheet(self, wb, sector_name: str, df_sector: pd.DataFrame):
        """构建单个板块明细 Sheet：先子概念排名，再按子概念分组列出个股明细"""
        safe_name = self._sanitize_sheet_title(sector_name)
        ws = wb.create_sheet(title=safe_name)

        # ========== 第一部分：子概念排名（含各维度打分次数）==========
        title_row = 1
        self.styler.merge_and_title(ws, title_row, 1, 16, f"=== {sector_name} 子概念排名（按平均分）===")

        # 计算子概念排名（按 子概念 分组，即 jiuyan_concept 按 - 拆分后的前半部分）
        # P1: 涨停统计统一用 market_data 已判定的"涨停"字段，删 >=9.9 重算
        zt_field = "涨停" if "涨停" in df_sector.columns else None
        if zt_field:
            sub_stats = df_sector.groupby("子概念").agg({
                "总得分": ["mean", "max", "count"],
                zt_field: "sum",
            }).reset_index()
            sub_stats.columns = ["子概念", "平均分", "最高分", "股票数量", "涨停数"]
        else:
            sub_stats = df_sector.groupby("子概念").agg({
                "总得分": ["mean", "max", "count"],
            }).reset_index()
            sub_stats.columns = ["子概念", "平均分", "最高分", "股票数量"]
            sub_stats["涨停数"] = 0
        sub_stats = sub_stats.sort_values("平均分", ascending=False).reset_index(drop=True)
        sub_stats["排名"] = range(1, len(sub_stats) + 1)
        sub_stats = sub_stats[["排名", "子概念", "股票数量", "平均分", "最高分", "涨停数"]]

        # 计算每个子概念在各维度上的打分次数（显示D1,D2,D4,D5,D6,D7,D8,D9,大成交额）
        dim_cols = ["D1强势形态且新高", "D2强势形态", "D4首板资金池", "D5潜在突破10日",
                    "D6潜在突破5日", "D7持续性", "D8情绪分数", "D9活跃程度", "大成交额额外加分"]
        dim_counts = {}
        for sub_concept in sub_stats["子概念"]:
            sub_df = df_sector[df_sector["子概念"] == sub_concept]
            counts = {}
            for dim in dim_cols:
                counts[dim] = int((sub_df[dim] > 0).sum()) if dim in sub_df.columns else 0
            dim_counts[sub_concept] = counts

        # 写入子概念排名表头（浅蓝色背景）
        header_row = 2
        headers = ["排名", "子概念", "股票数量", "平均分", "最高分", "涨停数",
                   "D1", "D2", "D4", "D5", "D6", "D7", "D8", "D9", "大成交额"]
        light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=c_idx, value=h)
            cell.fill = light_blue
            cell.font = self.styler.style.header_font_style
            cell.alignment = self.styler.style.header_alignment
            cell.border = self.styler.style.thin_border
        current_row = 3

        for _, row in sub_stats.iterrows():
            sub_concept = str(row["子概念"])
            ws.cell(row=current_row, column=1, value=int(row["排名"]))
            ws.cell(row=current_row, column=2, value=sub_concept)
            ws.cell(row=current_row, column=3, value=int(row["股票数量"]))
            ws.cell(row=current_row, column=4, value=round(float(row["平均分"]), 2))
            ws.cell(row=current_row, column=5, value=round(float(row["最高分"]), 2))
            ws.cell(row=current_row, column=6, value=int(row["涨停数"]))
            # 写入各维度打分次数
            counts = dim_counts.get(sub_concept, {})
            for c_idx, dim in enumerate(dim_cols, 7):
                ws.cell(row=current_row, column=c_idx, value=counts.get(dim, 0))
            # 应用数据样式，前排(D1/D2)和后排(D5/D6)区分颜色
            for c_idx in range(1, 16):
                self.styler.style.apply(ws, current_row, c_idx, "data")
            # 前排维度(D1/D2)用浅红色背景
            front_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
            for c_idx in [7, 8]:  # D1, D2
                ws.cell(row=current_row, column=c_idx).fill = front_fill
            # 后排维度(D5/D6)用浅蓝色背景
            back_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            for c_idx in [10, 11]:  # D5, D6
                ws.cell(row=current_row, column=c_idx).fill = back_fill
            current_row += 1

        # ========== 合计行 ==========
        total_row = current_row
        ws.cell(row=total_row, column=1, value="")
        ws.cell(row=total_row, column=2, value="合计")
        ws.cell(row=total_row, column=3, value=int(sub_stats["股票数量"].sum()))
        ws.cell(row=total_row, column=4, value="")
        ws.cell(row=total_row, column=5, value="")  # 最高分不需要求和
        ws.cell(row=total_row, column=6, value=int(sub_stats["涨停数"].sum()))
        # 维度求和
        for c_idx, dim in enumerate(dim_cols, 7):
            total = sum(dim_counts.get(sc, {}).get(dim, 0) for sc in sub_stats["子概念"])
            ws.cell(row=total_row, column=c_idx, value=total)
        # 合计行样式：前排(D1/D2)浅红色，后排(D5/D6)浅蓝色，其他浅黄色
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        front_total_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        back_total_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        for c_idx in range(1, 16):
            cell = ws.cell(row=total_row, column=c_idx)
            if c_idx in [7, 8]:      # D1, D2 - 前排
                cell.fill = front_total_fill
            elif c_idx in [10, 11]:   # D5, D6 - 后排
                cell.fill = back_total_fill
            else:
                cell.fill = total_fill
            cell.font = Font(name="微软雅黑", bold=True, color="000000", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.styler.style.thin_border
        current_row += 1

        # 空行分隔
        current_row += 2

        # ========== 第二部分：按子概念分组列出个股明细 ==========
        self.styler.merge_and_title(ws, current_row, 1, 16, f"=== {sector_name} 个股明细（按子概念分组，含 D1-D9+大成交额 打分）===")
        current_row += 1

        # 个股明细表头（浅绿色背景）
        detail_headers = ["排名", "代码", "名称", "所属板块", "总得分",
                          "涨跌幅", "成交额(亿)", "D1强势形态且新高", "D2强势形态",
                          "D4首板资金池", "D5潜在突破10日", "D6潜在突破5日",
                          "D7持续性", "D8情绪分数", "D9活跃程度", "大成交额额外加分"]
        light_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for c_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill = light_green
            cell.font = self.styler.style.header_font_style
            cell.alignment = self.styler.style.header_alignment
            cell.border = self.styler.style.thin_border
        current_row += 1

        # 按子概念分组写入个股（每个子概念下按总得分降序）
        for sub_concept in sub_stats["子概念"]:
            # 子概念分隔行（浅黄色背景）
            sub_df = df_sector[df_sector["子概念"] == sub_concept].copy()
            if sub_df.empty:
                continue
            sub_df = sub_df.sort_values("总得分", ascending=False).reset_index(drop=True)

            # 子概念分隔行
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
            cell = ws.cell(row=current_row, column=1, value=f"--- {sub_concept} ---")
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(name="微软雅黑", bold=True, color="000000", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = self.styler.style.thin_border
            current_row += 1

            for idx, row in sub_df.iterrows():
                ws.cell(row=current_row, column=1, value=idx + 1)
                ws.cell(row=current_row, column=2, value=str(row.get("代码", "")))
                ws.cell(row=current_row, column=3, value=str(row.get("名称", "")))
                ws.cell(row=current_row, column=4, value=str(row.get("所属板块", "")))
                ws.cell(row=current_row, column=5, value=int(row.get("总得分", 0)))
                ws.cell(row=current_row, column=6, value=row.get("涨跌幅", 0))
                amount = row.get("成交额", 0) or 0
                ws.cell(row=current_row, column=7, value=round(amount / 100000000, 2) if amount else 0)
                ws.cell(row=current_row, column=8, value=int(row.get("D1强势形态且新高", 0)))
                ws.cell(row=current_row, column=9, value=int(row.get("D2强势形态", 0)))
                ws.cell(row=current_row, column=10, value=int(row.get("D4首板资金池", 0)))
                ws.cell(row=current_row, column=11, value=int(row.get("D5潜在突破10日", 0)))
                ws.cell(row=current_row, column=12, value=int(row.get("D6潜在突破5日", 0)))
                ws.cell(row=current_row, column=13, value=int(row.get("D7持续性", 0)))
                ws.cell(row=current_row, column=14, value=int(row.get("D8情绪分数", 0)))
                ws.cell(row=current_row, column=15, value=int(row.get("D9活跃程度", 0)))
                ws.cell(row=current_row, column=16, value=int(row.get("大成交额额外加分", 0)))
                # 应用数据样式
                for c_idx in range(1, 17):
                    self.styler.style.apply(ws, current_row, c_idx, "data")
                # P2: 涨停股浅红底高亮 + 高分股加粗
                is_limit_up = int(row.get("涨停", 0)) == 1
                score = int(row.get("总得分", 0))
                limit_up_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                if is_limit_up:
                    for c_idx in range(1, 17):
                        ws.cell(row=current_row, column=c_idx).fill = limit_up_fill
                if score >= 15:
                    for c_idx in range(1, 17):
                        cell = ws.cell(row=current_row, column=c_idx)
                        if cell.font:
                            cell.font = Font(
                                name=cell.font.name or "微软雅黑",
                                bold=True,
                                color=cell.font.color or "000000",
                                size=cell.font.size or 9
                            )
                current_row += 1

        self.styler.auto_width(ws)
        return ws

    def build(self, data: dict, output_path: str):
        """
        组装完整报告
        :param data: {
            "summary": pd.DataFrame,      # 概念总排名
            "detail": pd.DataFrame,       # 详细个股排名
            "top5": List[dict],           # TOP5 列表
            "sectors": Dict[str, pd.DataFrame],  # 各板块明细
            "stats": dict                 # 执行统计指标（可选）
        }
        :param output_path: 输出 Excel 路径
        """
        wb = Workbook()

        # 0. 说明页（打分规则 + 执行统计）
        stats = data.get("stats", {})
        self._build_info_sheet(wb, stats)

        # 1. 概念总排名
        if "summary" in data and not data["summary"].empty:
            self._build_summary_sheet(wb, data["summary"], stats)
        else:
            wb.active.title = self.sheets_cfg.get("summary", "1-概念总排名")
            wb.active.cell(row=1, column=1, value="暂无数据")

        # 2. 详细个股排名
        if "detail" in data and not data["detail"].empty:
            self._build_detail_sheet(wb, data["detail"])

        # 3. TOP5
        if "top5" in data:
            self._build_top5_sheet(wb, data["top5"], data.get("summary"))

        # 4. 各板块明细
        if "sectors" in data:
            for sector_name, df_sector in data["sectors"].items():
                self._build_sector_sheet(wb, sector_name, df_sector)

        # 保存
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        # 如果文件已存在，先删除（避免Excel已打开导致的PermissionError）
        if os.path.exists(output_path):
            os.remove(output_path)
        wb.save(output_path)
        print(f"[OK] 报告已生成: {output_path}")


class ReportBuilder:
    """
    报告构建编排器
    将数据加载、评分、排名、样式、组装串联起来
    """

    def __init__(self, config_path: str = None):
        if config_path and os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                self.config = json.load(f)
        else:
            self.config = ExcelReporter._load_config()
        self.reporter = ExcelReporter(self.config)
        self.top5_ranker = Top5Ranker(
            max_same_sector=self.config.get("ranking", {}).get("max_same_sector_in_top5", 2),
            min_diversity=self.config.get("ranking", {}).get("top5_sector_diversity_min", 3),
            top_count=self.config.get("ranking", {}).get("top5_count", 5)
        )
        self.sector_ranker = SectorRanker()
        self.detail_ranker = DetailRanker()

    def build_all(self, df_summary: pd.DataFrame, df_detail: pd.DataFrame,
                  top5_list: List[dict], sector_data: Dict[str, pd.DataFrame],
                  stats: dict = None, output_path: str = None):
        """
        构建完整报告
        """
        if output_path is None:
            from datetime import datetime
            date_str = datetime.now().strftime("%Y%m%d")
            output_dir = self.config.get("base_dir", ".") + "/" + self.config.get("output_dir", "output")
            output_path = os.path.join(output_dir, f"韭研概念打分报告_{date_str}.xlsx")

        data = {
            "summary": df_summary,
            "detail": df_detail,
            "top5": top5_list,
            "sectors": sector_data,
            "stats": stats or {}
        }
        self.reporter.build(data, output_path)
        return output_path


# --------------- 扩展接口示例（注释） ---------------
# class PdfReporter(ReporterBase):
#     """PDF 输出示例"""
#     def build(self, data, output_path):
#         from reportlab.pdfgen import canvas
#         c = canvas.Canvas(output_path)
#         c.drawString(100, 800, "韭研概念打分报告")
#         c.save()
#
# 使用方式：
# builder = ReportBuilder()
# builder.reporter = PdfReporter()  # 替换输出格式
