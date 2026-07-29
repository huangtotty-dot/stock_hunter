# -*- coding: utf-8 -*-
"""
Excel 样式模块
职责：统一表头颜色、对齐、字体、列宽
扩展接口：继承 StyleBase 可新增主题（如浅色主题、深色主题）
"""
from abc import ABC, abstractmethod
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional


class StyleBase(ABC):
    """样式基类"""

    @abstractmethod
    def apply(self, ws, row: int, col: int, cell_type: str = "data"):
        """
        对指定单元格应用样式
        :param ws: openpyxl Worksheet
        :param row: 行号（1-based）
        :param col: 列号（1-based）
        :param cell_type: 单元格类型：header/title/separator/data
        """
        pass


class V8Style(StyleBase):
    """
    v8 标准样式
    - 一级标题：深蓝 #4472C4 + 白字 + 加粗居中
    - 二级标题：同上
    - 数据表头：同上
    - 三级分隔：浅绿 #E2EFDA + 黑字 + 加粗
    - 数据单元格：无填充 + 默认黑字
    """

    HEADER_BG = "4472C4"
    HEADER_FONT = "FFFFFF"
    SEPARATOR_BG = "E2EFDA"
    TITLE_BG = "4472C4"
    TITLE_FONT = "FFFFFF"

    def __init__(self, header_bg: str = None, header_font: str = None):
        self.header_bg = header_bg or self.HEADER_BG
        self.header_font = header_font or self.HEADER_FONT

        self.header_fill = PatternFill(start_color=self.header_bg, end_color=self.header_bg, fill_type="solid")
        self.header_font_style = Font(name="微软雅黑", bold=True, color=self.header_font, size=10)
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.title_fill = PatternFill(start_color=self.TITLE_BG, end_color=self.TITLE_BG, fill_type="solid")
        self.title_font_style = Font(name="微软雅黑", bold=True, color=self.TITLE_FONT, size=11)
        self.title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.separator_fill = PatternFill(start_color=self.SEPARATOR_BG, end_color=self.SEPARATOR_BG, fill_type="solid")
        self.separator_font = Font(name="微软雅黑", bold=True, color="000000", size=10)
        self.separator_alignment = Alignment(horizontal="center", vertical="center")

        self.data_font = Font(name="微软雅黑", color="000000", size=9)
        self.data_alignment_center = Alignment(horizontal="center", vertical="center")
        self.data_alignment_left = Alignment(horizontal="left", vertical="center")

        self.thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0")
        )

    def apply(self, ws, row: int, col: int, cell_type: str = "data"):
        cell = ws.cell(row=row, column=col)
        if cell_type == "header":
            cell.fill = self.header_fill
            cell.font = self.header_font_style
            cell.alignment = self.header_alignment
        elif cell_type == "title":
            cell.fill = self.title_fill
            cell.font = self.title_font_style
            cell.alignment = self.title_alignment
        elif cell_type == "separator":
            cell.fill = self.separator_fill
            cell.font = self.separator_font
            cell.alignment = self.separator_alignment
        else:  # data
            cell.font = self.data_font
            cell.alignment = self.data_alignment_center
        cell.border = self.thin_border

    def apply_row(self, ws, row: int, col_start: int, col_end: int, cell_type: str = "data"):
        """对整行应用样式"""
        for col in range(col_start, col_end + 1):
            self.apply(ws, row, col, cell_type)

    def auto_width(self, ws, min_width: int = 8, max_width: int = 50):
        """自动调整列宽（P2: CJK字符按2宽度计算，max_width扩至50）"""
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = min_width
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        try:
                            s = str(cell.value)
                            cell_len = 0
                            for ch in s:
                                if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯':
                                    cell_len += 2  # CJK字符按2宽度
                                else:
                                    cell_len += 1
                            max_length = max(max_length, cell_len)
                        except:
                            pass
            adjusted_width = min(max_width, max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width


class ExcelStyler:
    """
    Excel 样式统一入口
    对外提供简洁的样式应用接口
    """

    def __init__(self, style: StyleBase = None):
        self.style = style or V8Style()

    def apply_header_style(self, ws, row: int, col_start: int = 1, col_end: Optional[int] = None):
        """对指定行应用表头样式（深蓝 + 白字）"""
        if col_end is None:
            col_end = ws.max_column
        self.style.apply_row(ws, row, col_start, col_end, "header")

    def apply_title_style(self, ws, row: int, col_start: int = 1, col_end: Optional[int] = None):
        """对标题行应用标题样式"""
        if col_end is None:
            col_end = ws.max_column
        self.style.apply_row(ws, row, col_start, col_end, "title")

    def apply_separator_style(self, ws, row: int, col_start: int = 1, col_end: Optional[int] = None):
        """对分隔行应用浅绿分隔样式"""
        if col_end is None:
            col_end = ws.max_column
        self.style.apply_row(ws, row, col_start, col_end, "separator")

    def apply_data_style(self, ws, row: int, col_start: int = 1, col_end: Optional[int] = None):
        """对数据行应用默认样式"""
        if col_end is None:
            col_end = ws.max_column
        self.style.apply_row(ws, row, col_start, col_end, "data")

    def auto_width(self, ws):
        """自动调整所有列宽"""
        self.style.auto_width(ws)

    def merge_and_title(self, ws, row: int, col_start: int, col_end: int, title: str):
        """合并单元格并设置标题样式"""
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start)
        cell.value = title
        self.style.apply(ws, row, col_start, "title")


# --------------- 扩展接口示例（注释） ---------------
# class LightThemeStyle(StyleBase):
#     """浅色主题"""
#     def apply(self, ws, row, col, cell_type="data"):
#         if cell_type == "header":
#             ws.cell(row=row, column=col).fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
#             ws.cell(row=row, column=col).font = Font(bold=True, color="000000")
#         # ...
#
# 使用方式：
# styler = ExcelStyler(style=LightThemeStyle())
